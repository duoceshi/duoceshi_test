#coding=utf-8
"""
===========================
Author:多测师_王sir
Time:2020-12-16 20:36
Wechat:xiaoshubass
website:www.duoceshi.cn
===========================
"""

'''
此模块就是用来读取Excel表格里面的接口用例的
读取Excel表格有2个模块
xlrd
openpyxl
在dos窗口进行安装
pip install xlrd
pip install openpyxl
'''

import openpyxl
from common.handle_path import *
import os

class Read_Excel(object):

    def __init__(self,filename,sheet_name):
        '''
        :param filename: 读的文件名称
        :param sheet_name: 读取Excel表格中的哪个sheet表单
        '''
        self.filename = filename
        self.sheet_name = sheet_name

    def open(self):
        '''
        封装一个打开Excel表格的工具方法
        :return:
        '''
        #通过openpyxl模块里面的load_workbook函数读取对应的Excel文档
        self.wb = openpyxl.load_workbook(self.filename)
        #通过self.wb对象指定对应的sheet页面
        self.sh = self.wb[self.sheet_name]

    def get_data(self):
        '''封装一个获取Excel表格中数据的工具方法'''
        #1.打开Excel工作簿
        self.open()
        #2.把所有Excel表格的数据放入到了一个列表当中
        datas = list(self.sh.rows)
        #3.拿到表头的数据==》把表头的数据通过列表解析式放入到一个列表中
        title = [i.value for i in datas[0]]
        cases = []  # 定义一个列表用来装所有的用例
        #4.拿到接口用例
        for j in datas[1:]:
            values = [c.value for c in j]
            case = dict(zip(title,values))   #把表头和接口的用例数据用zip函数进行打包变成字典
            cases.append(case)
        return cases

    def write_data(self,row=None, column=None, value=None):
        '''封装一个用例跑完之后在Excel表格写入数据的工具方法'''
        #打开Excel表格
        self.open()
        #写入数据
        self.sh.cell(row, column, value)
        #保存Excel表格
        self.wb.save(self.filename)

# case_file = os.path.join(data_path,'apicases.xlsx')
# read_excel = Read_Excel(case_file,'login')
# print(read_excel.get_data())
# read_excel.write_data(2,8,'未通过')


























